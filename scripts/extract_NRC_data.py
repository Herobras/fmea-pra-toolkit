import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

xlsx_path = Path("seed_data/ParameterEstimates2020.xlsx")

wb = load_workbook(
    xlsx_path,
    read_only=True,   # más rápido y usa menos memoria
    data_only=True,   # lee valores calculados, no fórmulas
)

ws = wb["Component Reliability"]

rows = []
last_group = None
last_component_type = None

for row in ws.iter_rows(min_row = 6, max_row = 311, max_col = 20, values_only = True):
    group = row[0]

    #Solucionar celdas mergeadas
    if group is not None:
        last_group = group
    else:
        group = last_group

    group = str(group)

    component_type = row[1]
    if component_type is not None:
        last_component_type = component_type
    else:
        component_type = last_component_type

    component_type = str(component_type)

    failure_mode = str(row[2])
    failure_description = str(row[3])

    rows.append({"group": group, "component_type": component_type, "failure_mode": failure_mode, "failure_description": failure_description})

df = pd.DataFrame(rows)
wb.close()

df.to_csv("src/fmea_pra_toolkit/data/NUREG_CR_6928_2020.csv", index=False)